import clr
clr.AddReference("System.Drawing")
clr.AddReference("System.Windows.Forms")
from datetime import datetime
#from os import system

from connect import *
from openpyxl import Workbook
from openpyxl.styles import numbers
from System.Drawing import *
from System.Windows.Forms import *


def get_tx_technique(bs):
    # Helper function that returns the treatment technique for the given beam set.
    # SMLC, VMAT, DMLC, 3D-CRT, conformal arc, or applicator and cutout
    # Return "?" if treatment technique cannot be determined
    # Code modified from RS support

    if bs.Modality == "Photons":
        if bs.PlanGenerationTechnique == "Imrt":
            if bs.DeliveryTechnique == "SMLC":
                return "SMLC"
            if bs.DeliveryTechnique == "DynamicArc":
                if bs.Prescription is not None and bs.FractionationPattern is not None:
                    fx = beam_set.FractionationPattern.NumberOfFractions
                    rx = bs.Prescription.PrimaryDosePrescription.DoseValue / fx
                    if rx >= 600:
                        if fx.NumberOfFractions in [1, 3]:
                            return "SRS"
                        if fx.NumberOfFractions == 5:
                            return "SBRT"
                        if fx.NumberOfFractions in [6] + list(range(8, 16)):
                            return "SABR"
                return "VMAT"
            if bs.DeliveryTechnique == "DMLC":
                return "DMLC"
        elif bs.PlanGenerationTechnique == "Conformal":
            if bs.DeliveryTechnique == "SMLC":
                # return "SMLC" # Changed from "Conformal". Failing with forward plans.
                return "Conformal"
                # return "3D-CRT"
            if bs.DeliveryTechnique == "Arc":
                return "Conformal Arc"
    elif bs.Modality == "Electrons":
        if bs.PlanGenerationTechnique == "Conformal":
            if bs.DeliveryTechnique == "SMLC":
                return "ApplicatorAndCutout"
    return "[Unknown]"


class ListPatientsForm(Form):
    """Form that allows the user to specify filters for patient selection
    """

    def __init__(self):
        self.Text = "List Patients"
        self.FormBorderStyle = FormBorderStyle.FixedToolWindow  # User cannot resize form

        # Adapt form size to contents
        self.AutoSize = True
        self.AutoSizeMode = AutoSizeMode.GrowAndShrink
        self.MinimumSize = Size(TextRenderer.MeasureText(self.Text, SystemFonts.CaptionFont).Width + 100, 0)
        
        self.y = 15  # Vertical coordinate of next control

        # Script description/instructions
        l = Label()
        l.AutoSize = True
        l.Location = Point(15, self.y)
        text = "MRNs of patients meeting all of the following filter criteria will be written to a file in 'T:\Physics\Scripts\Output Files\ListPatients'.\n"
        text += "The filename includes a timestamp.\n"
        text += "Check the filters that you want to apply."
        l.Text = text
        self.Controls.Add(l)
        self.y += l.Height + 15

        # Number of patients to output
        l = Label()
        l.AutoSize = True
        l.Location = Point(15, self.y)
        l.Text = "Max # of MRN(s) to write:"
        self.Controls.Add(l)

        self.num_pts_tb = TextBox()
        self.num_pts_tb.Location = Point(20 + l.Width, self.y)
        self.num_pts_tb.Text = "[All]"
        self.num_pts_tb.Width = 25
        self.num_pts_tb.TextChanged += self.set_ok_enabled
        self.Controls.Add(self.num_pts_tb)
        self.y += 50

        # "Select all filters" checkbox
        self.select_all = CheckBox()
        self.select_all.IsThreeState = True
        self.select_all.CheckState = self.select_all.Tag = CheckState.Unchecked  # By default, no filters are applied
        self.select_all.Click += self.select_all_clicked
        self.select_all.AutoSize = True
        self.select_all.Location = Point(15, self.y)
        self.select_all.Text = "Select all filters"
        self.Controls.Add(self.select_all)
        self.y += self.select_all.Height + 10

        self.filter_cbs, self.filter_gbs = [], []  # Lists of checkboxes and their corresponding groupboxes, for each filter

        # Keywords filter
        instrs = "The following are searched for the keywords (case insensitive):\n    -  "
        instrs += "\n    -  ".join(["Case names, body sites, comments, and diagnoses", "Exam names", "Plan names and comments", "Beam set names and comments", "Beam names and descriptions", "DSP names", "Rx descriptions and structure names"])
        instrs += "\nEnter one keyword per line:"
        gb, gb_y = self.add_filter("Keywords", instrs)

        self.keyword_tb = TextBox()
        self.keyword_tb.AutoSize = True
        self.keyword_tb.Location = Point(15, gb_y)
        self.keyword_tb.Width = 50
        self.keyword_tb.MinimumSize = Size(100, 75)
        self.keyword_tb.Multiline = True  # Accepts more than one line
        #self.keyword_tb.ShortcutsEnabled = True
        self.keyword_tb.TextChanged += self.set_warning_label  # Empty vs. filled textbox influences whether any filters are actually applied
        gb.Controls.Add(self.keyword_tb)
        self.y += gb.Height + 10

        # Sex filter
        gb = self.add_checkboxes("Sex", ["Male", "Female", "Other"])
        self.y += gb.Height + 10

        # Patient Position filter
        gb = self.add_checkboxes("Patient Position", ["FFS", "HFP", "HFS"])
        self.y += gb.Height + 10

        # Tx technique filter
        gb = self.add_checkboxes("Treatment Technique", ["Applicator and cutout", "Conformal", "Conformal Arc", "DMLC", "SABR", "SBRT", "SRS", "SMLC", "VMAT", "[Unknown]"])
        self.y += gb.Height + 15

        # Warning label
        # Visible if no filters will be applied
        self.warning_lbl = Label()
        self.warning_lbl.AutoSize = True
        self.warning_lbl.Location = Point(15, self.y)
        self.warning_lbl.Text = "No filters are applied.\nAll MRNs will be written to the file."
        self.Controls.Add(self.warning_lbl)
        self.y += self.warning_lbl.Height

        # "OK" button
        self.ok = Button()
        self.ok.Location = Point(self.ClientSize.Width - 50, self.y)  # Right align
        self.ok.Text = "OK"
        self.ok.Click += self.ok_clicked
        self.Controls.Add(self.ok)
        self.ShowDialog()  # Launch window

    def add_filter(self, name, instrs=None):
        # Helper method that adds a checkbox and groupbox for a filter
        # `name`: Title of the filter
        # `instrs`: Instructions to display to the user
        # Return a 2-tuple: the groupbox and the y-coordinate for the next control in the groupbox

        # Checkbox
        cb = CheckBox()
        cb.Click += self.item_cb_clicked
        cb.AutoSize = True
        cb.Location = Point(30, self.y)
        self.Controls.Add(cb)

        ## Groupbox
        gb = GroupBox()
        gb.Enabled = False
        gb.AutoSize = True
        gb.AutoSizeMode = AutoSizeMode.GrowAndShrink
        gb.MinimumSize = Size(TextRenderer.MeasureText(name, gb.Font).Width + 50, 0)
        gb.Location = Point(50, self.y)
        gb.Text = "{}:".format(name)
        self.Controls.Add(gb)

        gb_y = 15  # Vertical coordinate of next control in groupbox

        # Filter description/instructions
        if instrs is not None:
            l = Label()
            l.AutoSize = True
            l.Location = Point(15, gb_y)
            l.Text = instrs
            gb.Controls.Add(l)
            gb_y += l.Height

        # Add new checkbox and groupbox to lists
        self.filter_cbs.append(cb)
        self.filter_gbs.append(gb)

        return gb, gb_y

    def add_checkboxes(self, name, items, instrs=None):
        # Helper method that adds a filter that is checkboxes
        # `name`: Title of the filter
        # `items`: List of checkbox options
        # `instrs`: Instructions to display to the user

        gb, gb_y = self.add_filter(name, instrs)  # Add checkbox and groupbox
        
        # "Select all" checkbox
        cb = CheckBox()
        cb.IsThreeState = True
        cb.CheckState = cb.Tag = CheckState.Unchecked  # By default, nothing checked
        cb.Click += self.select_all_clicked
        cb.AutoSize = True
        cb.Location = Point(15, gb_y)
        cb.Text = "Select all"
        gb.Controls.Add(cb)
        gb_y += cb.Height

        # Item checkboxes
        for item in items:
            cb = CheckBox()
            cb.Checked = False  # By default, nothing checked
            cb.AutoSize = True
            cb.Location = Point(30, gb_y)
            cb.Text = item
            cb.Click += self.item_cb_clicked
            gb.Controls.Add(cb)
            gb_y += cb.Height
        return gb

    def select_all_clicked(self, sender, event):
        # Helper method for clicking a "Select all" checkbox
        # Set checkstate of options checkboxes
        # If the checkbox is "Select all filters", enable or disable the corresponding groupboxes

        checked = sender.Tag == CheckState.Checked  # Checkbox tag tracks previous checkstate
        if checked:  # "Select all" is now checked -> uncheck
            sender.CheckState = sender.Tag = CheckState.Unchecked
        else:  # "Select all" is now unchecked or indeterminate -> check
            sender.CheckState = sender.Tag = CheckState.Checked

        # Set options checkstates
        parent = sender.Parent
        if isinstance(parent, Form):  # It's a filter checkbox
            for i, cb in enumerate(self.filter_cbs):
                gb = self.filter_gbs[i]
                cb.Checked = gb.Enabled = not checked  # Enable groupbox only if checkbox is checked
        else:  # It's a checkbox within a filter
            item_cbs = [ctrl for ctrl in parent.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is None]  # All checkboxes in the groupbox that are not "select all"
            for cb in item_cbs:
                cb.Checked = not checked  # Toggle checkstate

    def item_cb_clicked(self, sender, event):
        # Helper method for clicking a checkbox that is not "Select all"
        # Set corresponding "Select all" checkstate

        parent = sender.Parent
        select_all = [ctrl for ctrl in parent.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is not None][0]  # "Select all" checkbox in this group
        if isinstance(parent, Form):  # A filter checkbox was clicked
            gb = self.filter_gbs[self.filter_cbs.index(sender)]  # Corresponding groupbox
            gb.Enabled = sender.Checked  # Enable groupbox only if checkbox is now checked

        item_cbs_checked = [ctrl.Checked for ctrl in parent.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is None]  # All checkboxes in the groupbox that are not "Select all"
        if all(item_cbs_checked):  # All checkboxes are checked
            select_all.CheckState = select_all.Tag = CheckState.Checked  # Check "Select all"
        elif any(item_cbs_checked):  # Some checkboxes are checked, some unchecked
            select_all.CheckState = select_all.Tag = CheckState.Indeterminate  # "Select all" checkstate is indeterminate
        else:  # All checkboxes are unchecked
            select_all.CheckState = select_all.Tag = CheckState.Unchecked  # Uncheck "Select all"
        self.set_warning_label()  # Warning label visibility depends on whether any filters are applied

    def set_warning_label(self, *args):
        # Helper method that makes the warning label visible if no filters are applied, invisible otherwise
        # Effectively, a filter is not applied if the checkbox is unchecked (obviously), or if no keywords are provided (for keywords filter) or all checkboxes are checked (for other filters)
        # *args instead of `sender` and `event` because this method can also be called not an event handler

        vis = True  # Assume warning label should be visible
        for i, cb in enumerate(self.filter_cbs):
            if cb.CheckState == CheckState.Checked:  # Filter is checked, so if there's any meaningful input, it is applied
                gb = self.filter_gbs[i]  # Corresponding groupbox
                select_all = [ctrl for ctrl in gb.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is not None]  # "Select all" checkboxes in that groupbox (not present if it's the keywords groupbox)
                if select_all:  # It's a checkboxes groupbox
                    select_all = select_all[0]  # The "Select all" checkbox in the group
                    if select_all.CheckState == CheckState.Indeterminate:  # Indeterminate checkstate means some checkboxes arechecked and some unchecked, so there is meaningful input
                        vis = False  # No warning necessary since a filter is applied
                        break
                elif self.keyword_tb.Text.strip() != "":  # It's the keywords groupbox. If there's nothing but whitespace in the keywords textbox, the keyword filter is effectively not applied
                    vis = False  # There is text, so no warning necessary
                    break
        self.warning_lbl.Visible = vis  # Show/hide

    def set_ok_enabled(self, sender, event):
        # Helper method that enables/disables the "OK" button
        # Enable "OK" button only if the number of MRN(s) is set to "[All]" or a positive integer

        num_pts = self.num_pts_tb.Text
        try:
            num_pts = int(num_pts)
            self.ok.Enabled = num_pts > 0
        except:
            self.ok.Enabled = num_pts == "[All]"
        
        if self.ok.Enabled:
            self.num_pts_tb.BackColor = Color.White
        else:
            self.num_pts_tb.BackColor = Color.Red

    def ok_clicked(self, sender, event):
        # Event handler for clicking the "OK" button
        # Create `values` attribute to hold user-set filters

        self.max_num_pts = float("Inf") if self.num_pts_tb.Text == "[All]" else int(self.num_pts_tb.Text)
        self.values = {}  # keyword name : values
        for i, cb in enumerate(self.filter_cbs):  # Iterate over all filter checkboxes
            gb = self.filter_gbs[i]  # Corresponding groupbox
            name = gb.Text[:-1]  # Remove colon at end of groupbox text to get keyword name
            if cb.Checked:  # User wants to use the filter
                cbs = [ctrl for ctrl in gb.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is None]  # Non-"Select all" checkboxes
                if cbs:  # There are checkboxes, so this is not the keywords checkbox
                    self.values[name] = [cb.Text for cb in cbs if cb.Checked]  # Text of all checked checkboxes
                else:  # It's the keywords filter
                    self.values[name] = [keyword.strip().lower() for keyword in self.keyword_tb.Text.split("\n")]  # Each non-blank line in the keywords textboc
            else:  # User does not want to use the filter
                self.values[name] = None
        self.DialogResult = DialogResult.OK


def list_patients():
    """Write a CSV file with the MRNs of RayStation patients matching all the user-selected filters

    Filters:
    - Keywords: Any of the provided keywords is part of any of the following:
        * A case name, body site, comment, or dx
        * A plan name or comment
        * An exam name
        * A beam set name or comment
        * A beam name or description
        * A DSP name
        * An Rx description or structure name
    - Sex
    - Patient position: The patient position of an exam is as specified. If keywords are also provided, this exam must belong to the plan matching a keyword.
    - Treatment technique: The treatment delivery technique of a beam set. If keywords are also provided, this beam set must belong to the plan matching a keyword.

    Filename is the datetime and the number of matching patients
    Note that patients that are currently open will not be checked!
    If no patients are found, file contains one line: "No matching patients."
    """

    # Get filters from user
    form = ListPatientsForm()
    if form.DialogResult != DialogResult.OK:  # "OK" button was never clicked (user probably exited GUI)
        sys.exit()

    max_num_pts = form.max_num_pts
    keywords = form.values["Keywords"]
    sex = form.values["Sex"]
    pt_pos = form.values["Patient Position"]
    tx_techniques = form.values["Treatment Technique"]

    patient_db = get_current("PatientDB")

    matching_mrns = []  # MRNs to write to output file
    all_pts = patient_db.QueryPatientInfo(Filter={})  # Get all patient info in RS DB
    for pt in all_pts:  # Iterate over all patients in RS
        if len(matching_mrns) == max_num_pts:
            break

        try:
            pt = patient_db.LoadPatient(PatientInfo=pt)  # Load the patient
        except:  # Someone has the patient open
            continue
 
        # If sex filter applied and patient doesn't match, move on to next patient
        if sex is not None and pt.Gender not in sex:
            continue
        
        if keywords is not None:  # Keywords filter is applied
            names_to_chk = []  # Strings to check whether they contain any of the keywords
            for case in pt.Cases:
                for exam in case.Examinations:
                    if pt_pos is None or exam.PatientPosition in pt_pos:
                        names_to_chk.extend([case.BodySite, case.CaseName, case.Comments, case.Diagnosis])
                        names_to_chk.append(exam.Name)
                        for plan in case.TreatmentPlans:
                            if plan.GetStructureSet().OnExamination.Name.lower() == exam.Name.lower():
                                names_to_chk.extend([plan.Name, plan.Comments])
                                for bs in plan.BeamSets:
                                    if tx_techniques is None or get_tx_technique(bs) in tx_techniques:
                                        names_to_chk.extend([bs.Comment, bs.DicomPlanLabel])
                                        for b in bs.Beams:
                                            names_to_chk.extend([b.Description, b.Name])
                                        for dsp in bs.DoseSpecificationPoints:
                                            names_to_chk.append(dsp.Name)
                                        if bs.Prescription is not None:
                                            names_to_chk.append(bs.Prescription.Description)
                                            for rx in bs.Prescription.DosePrescriptions:
                                                if hasattr(rx, "OnStructure"):
                                                    names_to_chk.append(rx.OnStructure.Name)
           
            # Check each name for each keyword (case insensitive)
            keyword_match = False  # Assume no match
            for keyword in keywords:
                for name in set(names_to_chk):  # Remove duplicates from names list
                    if name is not None:  # E.g., beam description can be None
                        name = name.lower()
                        if keyword in name and (not keyword.startswith("pelvi") or "abd" not in name):  # Very special case: if searching for "pelvis", we don't want, e.g, "abdomen/pelvis"
                            keyword_match = True
                            break
            if not keyword_match:  # If no keywords present in any of the names, move on to next patient
                continue
        
        # Keywords filter is not applied, so exam with correct patient position doesn't have to be associated with a matching keyword
        elif pt_pos is not None and not any(exam.PatientPosition == pt_pos for case in pt.Cases for exam in case.Examinations):
            continue

        elif tx_techniques is not None and not any(get_tx_technique(bs) in tx_techniques for case in pt.Cases for plan in case.TreatmentPlans for bs in plan.BeamSets):
            continue
        
        matching_mrns.append(pt.PatientID)

    ## Write output file
    wb = Workbook()  # Create Excel data
    ws = wb.active
    
    # Format column as text, not number
    for row in ws:
        row[1].number_format = numbers.FORMAT_TEXT
    
    # Write data to file
    if matching_mrns:  # There are patients matching all criteria -> write in sorted order, one MRN per line
        for i, mrn in enumerate(sorted(matching_mrns)):
            ws.cell(i + 1, 1).value = mrn
    else:  # No matching patients, so say so in the first cell of the output file.
        ws.cell(1, 1).value = "No matching patients."

    # Write Excel data to file
    dt = datetime.now().strftime("%m-%d-%y %H_%M_%S")
    filename = "{} (1 Patient).xlsx".format(dt) if len(matching_mrns) == 1 else "{} ({} Patients).xlsx".format(dt, len(matching_mrns))
    wb.save(r"\\vs20filesvr01\groups\CANCER\Physics\Scripts\Output Files\ListPatients\{}".format(filename))

    # Open Excel file
    # No permissions to do this from RS script
    # Could always just write to a CSV file instead of XLSX, but then would lose ability to format as text (and thus the leading zeros in the MRN)
    # excel_path = r"\\Client\C$\Program Files (x86)\Microsoft Office\Office16\excel.exe"
    # system(r'START /B "{}" "{}"'.format(excel_path, filepath))
